import sys
from openpyxl import load_workbook

path = sys.argv[1]
wb = load_workbook(filename=path, read_only=True, data_only=True)
print('Sheets:', wb.sheetnames)

target = 'LUNES LABORES'
if target in wb.sheetnames:
    ws = wb[target]
    # read header (first non-empty row)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row and any(cell is not None for cell in row):
            print(f'Row {i}:', row)
            if i==1:
                headers = row
                print('Headers:', headers)
            if i>=5:
                break
else:
    # print first 3 rows of first sheet
    ws = wb[wb.sheetnames[0]]
    print('Target sheet not found; sample from first sheet:')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        print(f'Row {i}:', row)
        if i>=6:
            break
